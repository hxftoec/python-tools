import openpyxl

ipmap=[]

def readconf(config):
    file=open( config, 'r')
    for line in file:
        tmpline = line.strip().split()
        tmp = tmpline[0].strip(), tmpline[1].strip()
        ipmap.append( tmp ) 
    file.close()

readconf('config.txt')

wb = openpyxl.load_workbook('防火墙申请test.xlsx')
a_sheet = wb['Sheet1']

# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
#sheet = wb.active
# 获得最大列和最大行
#print(sheet.max_row)
#print(sheet.max_column)

for row in a_sheet.rows:
    for cell in row:
        for tmp in ipmap:
            oldip = tmp[0]
            cell.value=str(cell.value).replace( tmp[0], tmp[1] )
        #print(cell.value)

wb.save('new防火墙申请test.xlsx')
